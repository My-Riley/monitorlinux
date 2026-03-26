# module_admin/controller/student_import_controller.py
import io
import os
import re
from datetime import datetime
from zipfile import ZipFile

import pandas as pd
from fastapi import APIRouter, Depends, File, Query, Request, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl.drawing.image import Image as XLImage
from sqlalchemy import select, update
from sqlalchemy.ext.asyncio import AsyncSession

from config.env import UploadConfig
from config.get_db import get_db
from module_admin.entity.do.class_student_do import SysClassStudent
from module_admin.entity.vo.user_vo import CurrentUserModel
from module_admin.service.login_service import LoginService
from utils.log_util import logger
from utils.response_util import ResponseUtil

studentImportController = APIRouter(
    prefix="/system/student", dependencies=[Depends(LoginService.get_current_user)], tags=["学生导入"]
)

# 使用配置文件中的上传路径
UPLOAD_DIR = UploadConfig.UPLOAD_PATH


def delete_imported_face_image(face_image_url: str):
    """
    导入阶段的图片回滚删除：
    - 删除 /profile/... 对应的原始图片文件
    - 删除 upload_path/all_image 下的副本
    """
    if not face_image_url:
        return
    try:
        for img_path in str(face_image_url).split(","):
            img_path = img_path.strip()
            if not img_path:
                continue

            # /profile/xxx/yyy.jpg -> xxx/yyy.jpg
            relative_path = img_path.replace("/profile/", "")
            original_file = os.path.join(UploadConfig.UPLOAD_PATH, relative_path)
            if os.path.exists(original_file):
                os.remove(original_file)
                logger.info(f"导入回滚：已删除原始图片文件: {original_file}")

            filename = os.path.basename(img_path)
            all_image_path = os.path.join(UploadConfig.UPLOAD_PATH, "all_image", filename)
            if os.path.exists(all_image_path):
                os.remove(all_image_path)
                logger.info(f"导入回滚：已删除 all_image 图片副本: {all_image_path}")
    except Exception as e:
        logger.error(f"导入回滚：删除图片失败: {e}")


def save_image_to_file(
    image_data: bytes, student_id: str, student_name: str, old_student_id: str = None, old_student_name: str = None
) -> str:
    """将图片数据保存为文件并返回URL路径 - 统一存放到 all_image - 支持重命名"""
    try:
        all_image_dir = os.path.join(UPLOAD_DIR, "all_image")
        os.makedirs(all_image_dir, exist_ok=True)

        # 新文件名
        new_filename = f"{student_name}{student_id}.jpg"
        new_file_path = os.path.join(all_image_dir, new_filename)

        # === 删除旧图片 ===
        if old_student_id and old_student_name:
            old_filename = f"{old_student_name}{old_student_id}.jpg"
            old_file_path = os.path.join(all_image_dir, old_filename)

            # 只有当文件名变了（姓名或学号变了），才尝试删旧图
            if (old_student_id != student_id or old_student_name != student_name) and os.path.exists(old_file_path):
                try:
                    os.remove(old_file_path)
                    logger.info(f"旧图片已删除: {old_file_path}")
                except Exception as e:
                    logger.error(f"删除旧图片失败: {e}")

        # 直接保存（覆盖同名文件）
        with open(new_file_path, "wb") as f:
            f.write(image_data)

        url = f"/profile/all_image/{new_filename}"
        logger.info(f"图片已保存: {url}")
        return url

    except Exception as e:
        logger.error(f"保存图片失败: {str(e)}")
        return None


def extract_images_from_excel(file_bytes: bytes, df: pd.DataFrame) -> dict:
    """从Excel中提取图片并保存为文件（支持WPS和标准Excel格式）"""
    image_mapping = {}

    try:
        # 方法1: 尝试从标准Excel格式（openpyxl导出）中提取图片
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(file_bytes), data_only=False)
            worksheet = workbook.active  # 获取活动工作表

            # 获取工作表中的所有图片
            images = []
            if hasattr(worksheet, "_images") and worksheet._images:
                images = list(worksheet._images)

            logger.info(f"通过openpyxl找到 {len(images)} 张图片")

            if images:
                # 建立行号到学生信息的映射（DataFrame索引从0开始，Excel行号从2开始，第1行是标题）
                student_row_map = {}
                for df_idx, (idx, row) in enumerate(df.iterrows()):
                    student_id = str(row.get("学号", "")).strip()
                    if student_id and student_id != "nan" and student_id.isdigit():
                        # DataFrame行号（从0开始）+ 2 = Excel行号（第1行是标题，第2行开始是数据）
                        excel_row = df_idx + 2
                        student_row_map[excel_row] = {
                            "student_id": student_id,
                            "student_name": str(row.get("学生姓名", row.get("姓名", ""))).strip(),
                        }

                logger.info(f"建立了 {len(student_row_map)} 个学生的行号映射")

                # 按行号排序图片
                image_list = []
                for img_obj in images:
                    try:
                        # 尝试获取图片所在的行号
                        if hasattr(img_obj, "anchor") and img_obj.anchor:
                            anchor = img_obj.anchor
                            # openpyxl的图片定位方式可能不同，尝试多种方法
                            if hasattr(anchor, "_from"):
                                from_coord = anchor._from
                                if from_coord and hasattr(from_coord, "row"):
                                    img_row = from_coord.row + 1  # openpyxl行号从1开始
                                else:
                                    img_row = None
                            elif hasattr(anchor, "row"):
                                img_row = anchor.row + 1
                            else:
                                img_row = None

                            # 提取图片数据
                            if hasattr(img_obj, "_data"):
                                img_data = img_obj._data()
                            elif hasattr(img_obj, "ref"):
                                # 尝试从文件系统读取（不适用于内存中的数据）
                                img_data = None
                            else:
                                img_data = None

                            if img_data and img_row:
                                image_list.append({"row": img_row, "data": img_data, "img_obj": img_obj})
                    except Exception as e:
                        logger.warning(f"处理图片对象时出错: {str(e)}")
                        continue

                # 按行号排序图片
                image_list.sort(key=lambda x: x["row"])
                logger.info(f"成功提取 {len(image_list)} 张图片数据")

                # 匹配图片到学生
                for img_info in image_list:
                    img_row = img_info["row"]
                    img_data = img_info["data"]

                    if img_row in student_row_map:
                        student_info = student_row_map[img_row]
                        url = save_image_to_file(img_data, student_info["student_id"], student_info["student_name"])
                        if url:
                            image_mapping[student_info["student_id"]] = url
                            logger.info(
                                f"为学号 {student_info['student_id']} ({student_info['student_name']}) 保存图片 (行{img_row}): {url}"
                            )
                    else:
                        logger.warning(f"图片在行{img_row}，但找不到对应的学生数据")

                # 如果通过openpyxl成功提取到图片，直接返回
                if image_mapping:
                    logger.info(f"通过openpyxl共保存 {len(image_mapping)} 张图片文件")
                    return image_mapping
                else:
                    logger.warning("通过openpyxl提取到图片，但未能匹配到学生数据，尝试ZIP方法")
        except Exception as e:
            logger.warning(f"使用openpyxl提取图片失败，尝试其他方法: {str(e)}")
            import traceback

            logger.debug(traceback.format_exc())

        # 方法2: 从Excel ZIP中提取图片（支持WPS和标准Excel格式）
        with ZipFile(io.BytesIO(file_bytes), "r") as zip_file:
            all_files = zip_file.namelist()

            # 查找所有图片文件（WPS格式在media/，标准Excel在xl/media/）
            media_files = []
            for f in all_files:
                if "media" in f.lower() and f.lower().endswith((".png", ".jpg", ".jpeg", ".gif", ".bmp")):
                    media_files.append(f)

            media_files = sorted(media_files)
            logger.info(f"从ZIP中找到 {len(media_files)} 个媒体文件: {media_files}")

            if not media_files:
                logger.warning("未找到任何图片文件，可能Excel中没有嵌入图片")
                return {}

            # 建立学号到姓名的映射（按DataFrame顺序）
            student_list = []
            for index, row in df.iterrows():
                student_id = str(row.get("学号", "")).strip()
                if student_id and student_id != "nan" and student_id.isdigit():
                    student_name = str(row.get("学生姓名", row.get("姓名", ""))).strip()
                    student_list.append({"student_id": student_id, "student_name": student_name, "index": index})

            logger.info(f"找到 {len(student_list)} 个学生")

            if not student_list:
                logger.warning("未找到有效的学生数据")
                return {}

            # 图片匹配策略：
            # 1. 如果图片数量等于学生数量，按顺序一一对应
            # 2. 如果只有一张图片，所有学生共用
            # 3. 如果图片数量少于学生数量，前面的学生有图片，后面的没有
            if len(media_files) == 1 and len(student_list) > 1:
                # 所有学生共用一张图片
                logger.info(f"检测到所有学生共用一张图片: {media_files[0]}")
                img_data = zip_file.read(media_files[0])

                for student in student_list:
                    url = save_image_to_file(img_data, student["student_id"], student["student_name"])
                    if url:
                        image_mapping[student["student_id"]] = url
                        logger.info(f"为学号 {student['student_id']} ({student['student_name']}) 保存图片: {url}")
                else:
                    # 按顺序匹配图片到学生（假设Excel中图片顺序与数据行顺序一致）
                    for i, student in enumerate(student_list):
                        if i < len(media_files):
                            media_file = media_files[i]
                            img_data = zip_file.read(media_file)
                            url = save_image_to_file(img_data, student["student_id"], student["student_name"])
                            if url:
                                image_mapping[student["student_id"]] = url
                                logger.info(
                                    f"按顺序匹配: 学号 {student['student_id']} ({student['student_name']}) <- {media_file}"
                                )
                    else:
                        logger.warning(f"学号 {student['student_id']} 没有对应的图片（图片数量不足）")

        logger.info(f"共保存 {len(image_mapping)} 张图片文件")
        return image_mapping

    except Exception as e:
        logger.error(f"提取Excel图片失败: {str(e)}")
        import traceback

        traceback.print_exc()
        return {}


@studentImportController.post("/import")
async def upload_student_excel(
    request: Request,
    file: UploadFile = File(...),
    update_support: int = Query(default=0, alias="updateSupport"),
    query_db: AsyncSession = Depends(get_db),
    current_user: CurrentUserModel = Depends(LoginService.get_current_user),
):
    """导入学生Excel文件"""
    try:
        # 验证文件类型
        if not file.filename.endswith((".xlsx", ".xls")):
            return ResponseUtil.error(msg="只支持上传.xlsx或.xls格式的Excel文件")

        # 读取文件内容
        contents = await file.read()

        # 尝试读取Excel文件
        try:
            df = pd.read_excel(io.BytesIO(contents))
        except Exception as e:
            logger.error(f"读取Excel文件失败: {str(e)}")
            return ResponseUtil.error(msg=f"Excel文件格式错误或已损坏: {str(e)}")

        # 检查是否为空文件
        if df.empty:
            return ResponseUtil.error(msg="Excel文件内容为空，请检查文件")

        # 验证必需的列（支持"姓名"和"学生姓名"两种字段名，保持向后兼容）
        required_columns = ["学号"]
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return ResponseUtil.error(msg=f"Excel文件缺少必需的列: {', '.join(missing_columns)}")

        # 检查姓名字段（支持"姓名"或"学生姓名"）
        if "学生姓名" not in df.columns and "姓名" not in df.columns:
            return ResponseUtil.error(msg="Excel文件缺少必需的列: 姓名 或 学生姓名")

        # 提取Excel中的图片（支持WPS和标准Excel格式）并保存为文件
        logger.info("开始提取Excel中的图片...")
        image_mapping = extract_images_from_excel(contents, df)
        logger.info(f"图片提取完成，共保存 {len(image_mapping)} 张图片文件")

        success_count = 0
        failure_count = 0
        failure_msg = []
        image_count = 0

        for index, row in df.iterrows():
            try:
                # 提取数据并处理空值
                student_id = str(row.get("学号", "")).strip()
                if student_id.lower() == "nan" or student_id == "":
                    failure_msg.append(f"第{index + 2}行：学号为空")
                    failure_count += 1
                    continue

                # 支持"姓名"和"学生姓名"两种字段名（优先使用"学生姓名"）
                student_name = str(row.get("学生姓名", row.get("姓名", ""))).strip()
                if student_name.lower() == "nan" or student_name == "":
                    failure_msg.append(f"第{index + 2}行：姓名为空")
                    failure_count += 1
                    continue

                # 处理性别
                sex = str(row.get("性别", "0")).strip()
                if sex.lower() == "nan":
                    sex = "0"
                sex_map = {"男": "0", "女": "1", "M": "0", "F": "1", "male": "0", "female": "1"}
                sex = sex_map.get(sex, sex if sex in ["0", "1"] else "0")

                # 处理年龄
                age = row.get("年龄", None)
                if pd.isna(age):
                    age = None
                elif isinstance(age, (int, float)):
                    age = int(age)
                elif str(age).isdigit():
                    age = int(age)
                else:
                    age = None

                # 处理年级和班级
                grade = str(row.get("年级", "")).strip()
                if grade.lower() == "nan":
                    grade = ""

                if len(grade) > 10:
                    # 该行最终失败，回滚已保存的图片
                    face_image_for_cleanup = image_mapping.get(student_id, None)
                    if face_image_for_cleanup:
                        delete_imported_face_image(face_image_for_cleanup)
                        image_mapping.pop(student_id, None)
                    failure_msg.append(f"第{index + 2}行：年级字段过长（最多10字符），当前为 '{grade[:30]}...' ")
                    failure_count += 1
                    continue

                classes = str(row.get("班级", "")).strip()
                if classes.lower() == "nan":
                    classes = ""

                # 校验班级字段（必须是4位数字+“班”，总长5）
                if classes:
                    if len(classes) != 5 or not re.match(r"^\d{4}班$", classes):
                        # 该行最终失败，回滚已保存的图片
                        face_image_for_cleanup = image_mapping.get(student_id, None)
                        if face_image_for_cleanup:
                            delete_imported_face_image(face_image_for_cleanup)
                            image_mapping.pop(student_id, None)
                        failure_msg.append(f"第{index + 2}行：班级格式错误，应为4位数字加'班'，例如'2501班'")
                        failure_count += 1
                        continue

                # 获取对应的人脸图片URL（文件路径）
                face_image = image_mapping.get(student_id, None)
                if face_image:
                    image_count += 1

                # 验证学号格式
                if not student_id.isdigit():
                    # 该行最终失败，回滚已保存的图片
                    if face_image:
                        delete_imported_face_image(face_image)
                        image_mapping.pop(student_id, None)
                    failure_msg.append(f"第{index + 2}行：学号格式错误，必须为纯数字")
                    failure_count += 1
                    continue

                # 检查学号是否已存在
                existing_user_query = select(SysClassStudent).where(
                    SysClassStudent.student_id == student_id, SysClassStudent.del_flag == "0"
                )
                existing_result = await query_db.execute(existing_user_query)
                existing_user = existing_result.scalar()

                if existing_user:
                    if update_support == 1:
                        # 更新已存在的学生
                        update_data = {
                            "student_name": student_name,
                            "sex": sex,
                            "age": age,
                            "grade": grade,
                            "classes": classes,
                            "face_image": face_image,
                            "update_by": current_user.user.user_name,
                            "update_time": datetime.now(),
                        }
                        # 如果有图片，也更新图片
                        if face_image:
                            update_data["face_image"] = face_image

                        update_stmt = (
                            update(SysClassStudent)
                            .where(SysClassStudent.student_id == existing_user.student_id)
                            .values(**update_data)
                        )
                        await query_db.execute(update_stmt)
                        success_count += 1
                    else:
                        # 该行最终失败（不允许覆盖），回滚已保存的图片
                        if face_image:
                            delete_imported_face_image(face_image)
                            image_mapping.pop(student_id, None)
                        failure_msg.append(f"第{index + 2}行：学号{student_id}已存在")
                        failure_count += 1

                else:
                    # 新增
                    new_student = SysClassStudent(
                        student_id=int(student_id),
                        student_name=student_name,
                        sex=sex,
                        age=age,
                        grade=grade,
                        classes=classes,
                        face_image=face_image,
                        del_flag="0",
                        create_by=current_user.user.user_name,
                        create_time=datetime.now(),
                    )
                    query_db.add(new_student)
                    success_count += 1

            except Exception as e:
                # 异常导致该行失败，回滚已保存的图片
                try:
                    student_id_for_cleanup = str(row.get("学号", "")).strip()
                    if student_id_for_cleanup:
                        face_image_for_cleanup = image_mapping.get(student_id_for_cleanup, None)
                        if face_image_for_cleanup:
                            delete_imported_face_image(face_image_for_cleanup)
                            image_mapping.pop(student_id_for_cleanup, None)
                except Exception:
                    pass
                failure_msg.append(f"第{index + 2}行：{str(e)}")
                failure_count += 1
                logger.error(f"导入第{index + 2}行数据失败: {str(e)}")

        # 提交事务
        await query_db.commit()

        # 构建返回消息
        msg_parts = [f"成功导入{success_count}条数据"]
        if image_count > 0:
            msg_parts.append(f"，其中{image_count}条包含人脸图片")
        if failure_count > 0:
            msg_parts.append(f"，失败{failure_count}条")
            if failure_msg:
                msg_parts.append("<br/><br/>错误详情：<br/>")
                msg_parts.append("<br/>".join(failure_msg[:10]))
                if failure_count > 10:
                    msg_parts.append(f"<br/>...还有{failure_count - 10}条错误未显示")

        msg = "".join(msg_parts)
        logger.info(f"导入学生Excel完成：成功{success_count}条（含图片{image_count}条），失败{failure_count}条")

        return ResponseUtil.success(msg=msg)

    except Exception as e:
        await query_db.rollback()
        logger.error(f"导入学生Excel失败: {str(e)}")
        return ResponseUtil.error(msg=f"导入失败: {str(e)}")


@studentImportController.get("/importTemplate")
async def download_import_template(request: Request):
    """下载学生导入模板（包含示例人脸图片）"""
    output = io.BytesIO()

    # 模板字段顺序与导出格式保持一致
    template_data = {
        "年级": ["一年级", "二年级"],
        "班级": ["2501班", "2502班"],
        "学生姓名": ["学生A", "学生B"],
        "性别": ["男", "女"],
        "年龄": [15, 14],
        "学号": ["20250001", "20250002"],
        "人脸图片": ["", ""],  # 图片将通过openpyxl插入
    }
    df = pd.DataFrame(template_data)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="学生信息")

        # 设置列宽
        worksheet = writer.sheets["学生信息"]
        worksheet.column_dimensions["A"].width = 15  # 年级
        worksheet.column_dimensions["B"].width = 20  # 班级
        worksheet.column_dimensions["C"].width = 12  # 学生姓名
        worksheet.column_dimensions["D"].width = 8  # 性别
        worksheet.column_dimensions["E"].width = 8  # 年龄
        worksheet.column_dimensions["F"].width = 15  # 学号
        worksheet.column_dimensions["G"].width = 18  # 人脸图片

        # 设置标题行样式
        from openpyxl.styles import Alignment, Font

        header_font = Font(bold=True)

        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 插入示例人脸图片
        try:
            from openpyxl.drawing.image import Image as XLImage

            # 获取模板图片路径（新路径：backend/utils/template）
            template_dir = os.path.join(
                os.path.dirname(os.path.dirname(os.path.dirname(__file__))), "utils", "template"
            )
            face1_path = os.path.join(template_dir, "学生A202501.png")
            face2_path = os.path.join(template_dir, "学生B202502.png")

            # 设置行高以适应图片
            worksheet.row_dimensions[2].height = 80
            worksheet.row_dimensions[3].height = 80

            # 插入第一张图片（学生A）
            if os.path.exists(face1_path):
                img1 = XLImage(face1_path)
                img1.width = 80
                img1.height = 100
                worksheet.add_image(img1, "G2")
                logger.info(f"已插入模板图片1: {face1_path}")

            # 插入第二张图片（学生B）
            if os.path.exists(face2_path):
                img2 = XLImage(face2_path)
                img2.width = 80
                img2.height = 100
                worksheet.add_image(img2, "G3")
                logger.info(f"已插入模板图片2: {face2_path}")

        except Exception as e:
            logger.warning(f"插入模板图片失败: {str(e)}")

    output.seek(0)
    logger.info("下载学生导入模板成功")

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=student_import_template.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )


@studentImportController.get("/export")
async def export_student_list(
    request: Request,
    query_db: AsyncSession = Depends(get_db),
):
    """导出学生列表为Excel文件（包含人脸图片）"""

    # 查询所有有效学生（del_flag='0'）
    query = select(SysClassStudent).where(SysClassStudent.del_flag == "0").order_by(SysClassStudent.create_time.desc())

    result = await query_db.execute(query)
    students = result.scalars().all()

    if not students:
        return ResponseUtil.error(msg="暂无学生数据可导出")

    # 准备导出数据
    export_data = []
    image_paths = []  # 存储实际图片路径用于嵌入
    sex_map = {"0": "男", "1": "女", "2": "未知"}

    for student in students:
        # 性别转换
        sex_label = sex_map.get(student.sex, "未知")

        # 处理人脸图片路径
        face_image_url = student.face_image or ""
        face_image_path = None

        if face_image_url:
            # 只取第一个图片（如果有多个逗号分隔）
            first_url = face_image_url.split(",")[0].strip()
            if first_url.startswith("/profile/"):
                relative_path = first_url.replace("/profile/", "", 1)
                face_image_path = os.path.join(UploadConfig.UPLOAD_PATH, relative_path)
                if not os.path.exists(face_image_path):
                    logger.warning(f"导出时图片不存在: {face_image_path}")
                    face_image_path = None

        export_data.append(
            {
                "年级": student.grade or "",
                "班级": student.classes or "",
                "学生姓名": student.student_name or "",
                "性别": sex_label,
                "年龄": student.age if student.age is not None else "",
                "学号": str(student.student_id) if student.student_id else "",
                "人脸图片": face_image_url,  # 文本备份
                "创建时间": student.create_time.strftime("%Y-%m-%d %H:%M:%S") if student.create_time else "",
            }
        )

        image_paths.append(face_image_path)

    # 生成Excel
    output = io.BytesIO()
    df = pd.DataFrame(export_data)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="学生信息")
        worksheet = writer.sheets["学生信息"]

        # 设置列宽
        col_widths = {
            "A": 15,  # 年级
            "B": 20,  # 班级
            "C": 12,  # 学生姓名
            "D": 8,  # 性别
            "E": 8,  # 年龄
            "F": 15,  # 学号
            "G": 25,  # 人脸图片
            "H": 20,  # 创建时间
        }
        for col, width in col_widths.items():
            worksheet.column_dimensions[col].width = width

        # 插入图片（G列，从第2行开始）
        for idx, img_path in enumerate(image_paths, start=2):
            if img_path and os.path.exists(img_path):
                try:
                    img = XLImage(img_path)
                    # 自动缩放图片以适应单元格（最大80x80）
                    from PIL import Image as PILImage

                    with PILImage.open(img_path) as pil_img:
                        orig_w, orig_h = pil_img.size
                        scale = min(80 / orig_w, 80 / orig_h, 1.0)
                        new_w = int(orig_w * scale)
                        new_h = int(orig_h * scale)
                    img.width = new_w
                    img.height = new_h

                    cell = f"G{idx}"
                    worksheet.add_image(img, cell)
                    worksheet[cell].value = ""  # 清空文本
                    worksheet.row_dimensions[idx].height = 60  # 设置行高
                except Exception as e:
                    logger.error(f"插入学生图片失败 ({img_path}): {e}")

        # 标题样式
        from openpyxl.styles import Alignment, Font

        header_font = Font(bold=True)
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 内容垂直居中
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(vertical="center")

    output.seek(0)
    logger.info(f"学生列表导出成功，共 {len(students)} 条记录")

    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": "attachment; filename=student_list.xlsx",
            "Access-Control-Expose-Headers": "Content-Disposition",
        },
    )
